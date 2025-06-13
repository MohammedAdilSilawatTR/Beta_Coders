from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from pydantic import BaseModel
from typing import List, Dict, Any
import openpyxl
import json
from openai import OpenAI
import os # For API Key
from dotenv import load_dotenv
from datetime import datetime # Added for datetime conversion

# Attempt to import the categorization function
try:
    from agentic import get_batch_categories # UPDATED to get_batch_categories
except ImportError:
    print("WARN: agentic.py or get_batch_categories not found. Categorization endpoint will not work.")
    get_batch_categories = None # UPDATED to get_batch_categories

app = FastAPI()

workbook = None # Initialize workbook variable
final_processed_data_rows = [] # Initialize final processed data rows variable
header_mapping = {} # Initialize header mapping variable
actual_headers = [] # Initialize actual headers variable

# Initialize OpenAI client - Assumes OPENAI_API_KEY environment variable is set
# You might want to add more robust error handling for API key loading in a production app
try:
    load_dotenv()  # Load environment variables from .env file if it exists
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"),
                    base_url="https://litellm.int.thomsonreuters.com")
except TypeError:
    print("ERROR: OPENAI_API_KEY environment variable not set.")
    client = None # Or handle this more gracefully

PREDEFINED_COLUMNS = {"amount": "datatype is number", "transactionDate": "datatype is date", "transactionDescription": "datatype is string", "disallowableExpenses": "datatype is number"}
MAX_SAMPLE_ROWS = 5 # Number of sample data rows to send to LLM for each column

def convert_datetimes_to_string(obj):
    """
    Recursively convert datetime objects in nested lists/dictionaries to ISO format strings.
    """
    if isinstance(obj, list):
        return [convert_datetimes_to_string(i) for i in obj]
    elif isinstance(obj, dict):
        return {k: convert_datetimes_to_string(v) for k, v in obj.items()}
    elif isinstance(obj, datetime):
        return obj.isoformat()
    return obj

@app.post("/uploadfile/")
async def create_upload_file(
    file: UploadFile = File(...),
    business_description: str = Form("") 
):
    try:
        print(f"Received file: {file.filename}")
        print(f"Received Business Description: {business_description}")
        workbook = openpyxl.load_workbook(file.file)
        
        sheet = workbook.active
        
        non_empty_columns_indexes = [
            i for i, col in enumerate(sheet.iter_cols(values_only=True))
            if any(cell is not None for cell in col)
        ]

        actual_headers = []
        processed_data_rows = []
        header_row_found_excel_idx = -1 # 1-based index of the header row in Excel

        if not non_empty_columns_indexes:
            print("Warning: No non-empty columns found. Headers and data rows will be empty.")
        else:
            for current_excel_row_idx, row_tuple in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Extract values from the current row based on non_empty_columns_indexes
                current_row_selected_cols = []
                for col_idx in non_empty_columns_indexes:
                    if col_idx < len(row_tuple):
                        current_row_selected_cols.append(row_tuple[col_idx])
                    else:
                        current_row_selected_cols.append(None) 
                
                # Check if this row (after selecting relevant columns) has any content
                if any(cell_value is not None for cell_value in current_row_selected_cols):
                    if header_row_found_excel_idx == -1: # Header not yet found
                        actual_headers = [str(h) if h is not None else f"Unknown_Header_{i}" 
                                          for i, h in enumerate(current_row_selected_cols)]
                        header_row_found_excel_idx = current_excel_row_idx
                        print(f"Header row found at Excel row index: {header_row_found_excel_idx}")
                    else: # Header already found, this is a data row
                        processed_data_rows.append(current_row_selected_cols)
            
            if header_row_found_excel_idx == -1:
                print("Warning: No non-empty row found to be used as header. Headers and data will be empty.")

        print(f"Sheet Name: {sheet.title}")
        print(f"Actual Headers: {actual_headers}")
        # print(f"Processed Data Rows (first few): {processed_data_rows[:5]}") # Print first 5 for brevity
        if header_row_found_excel_idx != -1 and not processed_data_rows:
            print("Info: Header row found, but no subsequent data rows.")

        header_mapping = {}
        # Use actual_headers for the condition and further processing
        if client and actual_headers: 
            # Prepare sample data for LLM using actual_headers and processed_data_rows
            raw_sample_data_for_llm = {}
            for i, header_name in enumerate(actual_headers):
                column_data = [row[i] for row in processed_data_rows if i < len(row) and row[i] is not None]
                raw_sample_data_for_llm[header_name] = column_data[:MAX_SAMPLE_ROWS]
            
            # Convert datetimes in sample data before sending to LLM
            sample_data_for_llm = convert_datetimes_to_string(raw_sample_data_for_llm)
            print(f"Sample data for LLM (datetimes converted): {sample_data_for_llm}")

            # Construct prompt for OpenAI
            prompt_messages = [
                {
                    "role": "system",
                    "content": "You are an expert data mapping assistant. Your task is to map user-provided Excel column headers to a predefined list of standard column names. You will be given the user's headers, sample data from each of their columns, and the list of predefined standard columns. Remember amount will be greater than disallowableExpenses. Return your mapping as a JSON object where keys are the predefined standard columns and values are the matched user headers. If no good match is found for a user header, use null as its value."
                },
                {
                    "role": "user",
                    "content": f"""
User Headers:
{actual_headers}

Sample Data per User Header (first {MAX_SAMPLE_ROWS} non-null values):
{json.dumps(sample_data_for_llm, indent=2)}

Predefined Standard Columns:
{PREDEFINED_COLUMNS}

Please provide the mapping as a JSON object.
"""
                }
            ]

            try:
                print("Sending request to OpenAI...")
                chat_completion = client.chat.completions.create(
                    messages=prompt_messages,
                    model="openai/gpt-4o", # Or "gpt-4" or other preferred model
                    response_format={ "type": "json_object" } # Request JSON output
                )
                llm_response_content = chat_completion.choices[0].message.content
                print(f"LLM raw response: {llm_response_content}")
                if llm_response_content:
                    header_mapping = json.loads(llm_response_content)
                else:
                    # Consistent fallback: {predefined_column: None}
                    header_mapping = {predefined_col: None for predefined_col in PREDEFINED_COLUMNS}
                    print("LLM returned empty content. Using fallback mapping.")

            except Exception as llm_e:
                print(f"Error calling OpenAI or parsing response: {llm_e}")
                # Consistent fallback: {predefined_column: None}
                header_mapping = {predefined_col: None for predefined_col in PREDEFINED_COLUMNS}
        
        elif not client or not actual_headers: # Modified condition to also check actual_headers
            if not client:
                print("OpenAI client not initialized. Skipping LLM mapping.")
            if not actual_headers:
                print("No actual headers found. Skipping LLM mapping.")
            # Consistent fallback: {predefined_column: "Client/Headers Issue"}
            # Or simply None, but a string message might be more informative for this specific case
            header_mapping = {predefined_col: "OpenAI client not initialized or no headers" for predefined_col in PREDEFINED_COLUMNS}


        # Convert datetimes in processed_data_rows before returning
        final_processed_data_rows = convert_datetimes_to_string(processed_data_rows)

        return {
            "filename": file.filename,
            "headers": actual_headers, # Use actual_headers
            "non_empty_rows": final_processed_data_rows, # Use the correctly processed data rows
            "header_mapping": header_mapping
        }
    except Exception as e:
        print(f"Error in file processing: {e}")
        return {"error": str(e), "details": "Error during file processing or LLM interaction."}

# Pydantic models for the new endpoint
class TransactionData(BaseModel):
    transactionDescription: str | None = None # Making it optional in case it's missing
    # Include other fields from PREDEFINED_COLUMNS if needed by the agent or for context,
    # or allow any other fields by using Dict[str, Any]
    amount: float | int | None = None
    transactionDate: str | None = None
    disallowableExpenses: float | int | None = None
    # Allow other arbitrary fields that might be in the mapped_transactions
    class Config:
        extra = "allow"

class CategorizationRequest(BaseModel):
    business_description: str
    mapped_transactions: List[Dict[str, Any]] # Using Dict for flexibility from frontend

@app.post("/categorize-transactions/")
async def categorize_transactions_endpoint(request: CategorizationRequest):
    if not get_batch_categories: # UPDATED check
        raise HTTPException(status_code=501, detail="Categorization service is not available due to import error.")

    # Collect all valid transaction descriptions for batch processing
    descriptions_to_categorize = []
    original_indices_map = {} # To map description back to its original transaction object if needed, or just iterate later

    for i, transaction_data_dict in enumerate(request.mapped_transactions):
        desc = transaction_data_dict.get("transactionDescription")
        if isinstance(desc, str) and desc.strip():
            descriptions_to_categorize.append(desc)
            # If descriptions are not unique, this map might lose some original contexts.
            # Assuming for now that we will iterate through original transactions and use the map.
            # original_indices_map[desc] = i # Example if needed for direct update

    category_map = {}
    if descriptions_to_categorize:
        print(f"Sending batch of {len(descriptions_to_categorize)} descriptions for categorization with business: '{request.business_description}'")
        try:
            # Assuming get_batch_categories is a synchronous function as defined.
            # If it were async, this would need 'await'.
            category_map = get_batch_categories(
                business_query=request.business_description,
                descriptions_list=descriptions_to_categorize
            )
            print(f"Received category map: {category_map}")
        except Exception as e:
            print(f"Error calling get_batch_categories: {e}")
            # Fallback: mark all descriptions in this batch with an error
            category_map = {desc: f"Error during batch categorization: {str(e)}" for desc in descriptions_to_categorize}
    
    # Update each transaction in the original list using the category_map
    final_categorized_transactions = []
    for transaction_data_dict in request.mapped_transactions:
        desc = transaction_data_dict.get("transactionDescription")
        category_to_assign = "Missing or Invalid Description" # Default for invalid/missing descriptions

        if isinstance(desc, str) and desc.strip():
            # If description was valid, get its category from the map.
            # Default to "Uncategorized" if not found in map (e.g., LLM didn't return it or error).
            category_to_assign = category_map.get(desc, "Uncategorized")
        
        updated_transaction = transaction_data_dict.copy()
        updated_transaction["category"] = category_to_assign
        final_categorized_transactions.append(updated_transaction)

    return final_categorized_transactions
